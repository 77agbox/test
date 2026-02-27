import asyncio
import logging
import os
import re
from typing import List, Dict

import openpyxl
from aiogram import Bot, Dispatcher, types
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardRemove
from aiogram.utils.keyboard import ReplyKeyboardBuilder
from dotenv import load_dotenv

# ================= CONFIG =================

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID"))

logging.basicConfig(level=logging.INFO)

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())

# ================= DATA =================

def load_clubs(file="joined_clubs.xlsx") -> List[Dict]:
    if not os.path.exists(file):
        logging.warning("Файл joined_clubs.xlsx не найден")
        return []

    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    sheet = wb.active

    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}
        for h, v in zip(headers, row):
            record[h] = str(v).strip() if v else ""
        data.append(record)

    logging.info(f"Загружено кружков: {len(data)}")
    return data


CLUBS_DATA = load_clubs()

PACKAGE_MODULES = {
    "Картинг": [2200, 2100, 2000],
    "Симрейсинг": [1600, 1500, 1400],
    "Лазертаг": [1600, 1500, 1400],
}

MASTERCLASSES = [
    {
        "title": "Сумочка для телефона",
        "date": "04.03.2026",
        "time": "17:00",
        "price": 1500,
        "link": "https://t.me/dyutsvictory/3733",
    }
]

# ================= STATES =================

class ClubsForm(StatesGroup):
    address = State()
    age = State()
    direction = State()
    club = State()


class PackageForm(StatesGroup):
    people = State()
    activities = State()
    name = State()
    phone = State()


class MasterForm(StatesGroup):
    select = State()
    name = State()
    phone = State()


# ================= UTIL =================

def parse_age(age_str):
    if not age_str:
        return None, None
    age_str = age_str.lower()
    if "-" in age_str:
        try:
            a, b = age_str.split("-")
            return int(a), int(b)
        except:
            return None, None
    match = re.search(r"\d+", age_str)
    if match:
        return int(match.group()), 99
    return None, None


def main_menu():
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="Кружки", callback_data="m_clubs")],
        [InlineKeyboardButton(text="Пакетные туры", callback_data="m_package")],
        [InlineKeyboardButton(text="Мастер-классы", callback_data="m_master")]
    ])


def bottom_kb():
    builder = ReplyKeyboardBuilder()
    builder.button(text="Начать заново")
    builder.adjust(1)
    return builder.as_markup(resize_keyboard=True)


# ================= START =================

@dp.message(CommandStart())
async def start(message: types.Message):
    await message.answer("Главное меню:", reply_markup=bottom_kb())
    await message.answer("Выберите раздел:", reply_markup=main_menu())


@dp.message(lambda m: m.text == "Начать заново")
async def restart(message: types.Message, state: FSMContext):
    await state.clear()
    await start(message)

# ================= CLUBS =================

@dp.callback_query(lambda c: c.data == "m_clubs")
async def clubs_start(callback: types.CallbackQuery, state: FSMContext):
    addresses = sorted(list(set(c["Адрес предоставления услуги"] for c in CLUBS_DATA)))
    await state.update_data(addresses=addresses)

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=addr, callback_data=f"a_{i}")]
        for i, addr in enumerate(addresses)
    ])

    await state.set_state(ClubsForm.address)
    await callback.message.edit_text("Выберите адрес:", reply_markup=keyboard)
    await callback.answer()


@dp.callback_query(lambda c: c.data.startswith("a_"))
async def clubs_address(callback: types.CallbackQuery, state: FSMContext):
    idx = int(callback.data.split("_")[1])
    data = await state.get_data()
    address = data["addresses"][idx]

    await state.update_data(selected_address=address)
    await state.set_state(ClubsForm.age)

    await callback.message.edit_text("Введите возраст ребёнка:")
    await callback.answer()


@dp.message(ClubsForm.age)
async def clubs_age(message: types.Message, state: FSMContext):
    try:
        age = int(message.text)
    except:
        await message.answer("Введите число.")
        return

    data = await state.get_data()
    address = data["selected_address"]

    filtered = []
    for club in CLUBS_DATA:
        if club["Адрес предоставления услуги"] != address:
            continue
        min_a, max_a = parse_age(club["Возраст"])
        if min_a is None or (min_a <= age <= max_a):
            filtered.append(club)

    if not filtered:
        await message.answer("Подходящих кружков нет.", reply_markup=bottom_kb())
        await state.clear()
        return

    await state.update_data(filtered=filtered)

    directions = sorted(set(c["Наименование третьего уровня РБНДО"] for c in filtered))
    await state.update_data(directions=directions)

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=d, callback_data=f"d_{i}")]
        for i, d in enumerate(directions)
    ])

    await state.set_state(ClubsForm.direction)
    await message.answer("Выберите направление:", reply_markup=keyboard)


@dp.callback_query(lambda c: c.data.startswith("d_"))
async def clubs_direction(callback: types.CallbackQuery, state: FSMContext):
    idx = int(callback.data.split("_")[1])
    data = await state.get_data()
    direction = data["directions"][idx]

    clubs = [c for c in data["filtered"] if c["Наименование третьего уровня РБНДО"] == direction]
    await state.update_data(final=clubs)

    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(
            text=c["Наименование детского объединения"][:50],
            callback_data=f"c_{i}"
        )] for i, c in enumerate(clubs)
    ])

    await state.set_state(ClubsForm.club)
    await callback.message.edit_text("Выберите кружок:", reply_markup=keyboard)
    await callback.answer()


@dp.callback_query(lambda c: c.data.startswith("c_"))
async def clubs_show(callback: types.CallbackQuery, state: FSMContext):
    idx = int(callback.data.split("_")[1])
    data = await state.get_data()
    club = data["final"][idx]

    text = (
        f"<b>{club['Наименование детского объединения']}</b>\n\n"
        f"Возраст: {club['Возраст']}\n"
        f"Адрес: {club['Адрес предоставления услуги']}\n"
        f"Педагог: {club.get('Педагог','—')}\n\n"
        f"{club.get('Ссылка','')}"
    )

    await callback.message.edit_text(text, parse_mode="HTML", disable_web_page_preview=True)
    await callback.answer()

# ================= PACKAGE =================

PACKAGE_MODULES = {
    "Картинг": [2200, 2100, 2000],
    "Симрейсинг": [1600, 1500, 1400],
    "Практическая стрельба": [1600, 1500, 1400],
    "Лазертаг": [1600, 1500, 1400],
    "Керамика": [1600, 1500, 1400],
    "Мягкая игрушка": [1300, 1200, 1100],
}


class PackageForm(StatesGroup):
    people = State()
    activities = State()
    name = State()
    phone = State()


def activities_keyboard(selected=None):
    selected = selected or []
    builder = ReplyKeyboardBuilder()

    for name in PACKAGE_MODULES:
        text = f"{name} {'✅' if name in selected else ''}".strip()
        builder.button(text=text)

    # Кнопка "Готово" отдельной строкой
    builder.button(text="🟢 Готово")
    builder.adjust(2, 1)

    return builder.as_markup(resize_keyboard=True)


@dp.callback_query(lambda c: c.data == "m_package")
async def package_start(callback: types.CallbackQuery, state: FSMContext):
    await state.set_state(PackageForm.people)
    await callback.message.answer(
        "Сколько человек в группе?\n\nМинимум от 5 человек."
    )
    await callback.answer()


@dp.message(PackageForm.people)
async def package_people(message: types.Message, state: FSMContext):
    try:
        people = int(message.text)
        if people < 5:
            await message.answer("Минимальное количество — от 5 человек.")
            return
    except:
        await message.answer("Введите число.")
        return

    await state.update_data(people=people, selected=[])
    await state.set_state(PackageForm.activities)

    await message.answer(
        "Выберите от 1 до 3 активностей:",
        reply_markup=activities_keyboard(),
    )


@dp.message(PackageForm.activities)
async def package_activities(message: types.Message, state: FSMContext):
    text = message.text.replace(" ✅", "").strip()
    data = await state.get_data()
    selected = data.get("selected", [])

    if text == "🟢 Готово":
        if not 1 <= len(selected) <= 3:
            await message.answer("Выберите от 1 до 3 активностей.")
            return

        await state.set_state(PackageForm.name)
        await message.answer("Ваше имя:", reply_markup=ReplyKeyboardRemove())
        return

    if text in PACKAGE_MODULES:
        if text in selected:
            selected.remove(text)
        else:
            if len(selected) >= 3:
                await message.answer("Можно выбрать максимум 3 активности.")
                return
            selected.append(text)

        await state.update_data(selected=selected)

        await message.answer(
            f"Выбрано: {', '.join(selected) if selected else 'ничего'}",
            reply_markup=activities_keyboard(selected),
        )


@dp.message(PackageForm.name)
async def package_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text.strip())
    await state.set_state(PackageForm.phone)
    await message.answer("Телефон для связи:")


@dp.message(PackageForm.phone)
async def package_finish(message: types.Message, state: FSMContext):
    data = await state.get_data()

    people = data["people"]
    selected = data["selected"]
    name = data["name"]
    phone = message.text.strip()

    price_index = len(selected) - 1

    total = 0
    lines = []

    for act in selected:
        price_per_person = PACKAGE_MODULES[act][price_index]
        cost = price_per_person * people
        total += cost

        lines.append(
            f"{act}: {price_per_person} ₽ × {people} = {cost} ₽"
        )

    calc_text = "\n".join(lines)

    # === Сообщение админу ===
    await bot.send_message(
        ADMIN_ID,
        f"🛒 Новая заявка на пакетный тур\n\n"
        f"Имя: {name}\n"
        f"Телефон: {phone}\n"
        f"TG: @{message.from_user.username or 'нет username'}\n"
        f"TG ID: {message.from_user.id}\n\n"
        f"Количество человек: {people}\n"
        f"Активности: {', '.join(selected)}\n\n"
        f"{calc_text}\n\n"
        f"Итого: {total} ₽"
    )

    # === Сообщение клиенту ===
    await message.answer(
        f"✅ Ваша заявка принята!\n\n"
        f"Количество человек: {people}\n"
        f"Активности: {', '.join(selected)}\n\n"
        f"Расчёт стоимости:\n"
        f"{calc_text}\n\n"
        f"💰 Общая стоимость: {total} ₽\n\n"
        f"С вами скоро свяжется администратор для подтверждения.",
        reply_markup=bottom_kb(),
    )

    await state.clear()
# ================= MASTER =================

@dp.callback_query(lambda c: c.data == "m_master")
async def master_start(callback: types.CallbackQuery, state: FSMContext):
    keyboard = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=m["title"], callback_data=f"m_{i}")]
        for i, m in enumerate(MASTERCLASSES)
    ])

    await state.set_state(MasterForm.select)
    await callback.message.edit_text("Выберите мастер-класс:", reply_markup=keyboard)
    await callback.answer()


@dp.callback_query(lambda c: c.data.startswith("m_") and c.data != "m_master")
async def master_select(callback: types.CallbackQuery, state: FSMContext):
    idx = int(callback.data.split("_")[1])
    mc = MASTERCLASSES[idx]

    await state.update_data(mc=mc)
    await state.set_state(MasterForm.name)

    await callback.message.answer(f"{mc['title']}\nВведите имя:")
    await callback.answer()


@dp.message(MasterForm.name)
async def master_name(message: types.Message, state: FSMContext):
    await state.update_data(name=message.text)
    await state.set_state(MasterForm.phone)
    await message.answer("Телефон:")


@dp.message(MasterForm.phone)
async def master_finish(message: types.Message, state: FSMContext):
    data = await state.get_data()
    mc = data["mc"]

    await bot.send_message(
        ADMIN_ID,
        f"Запись на МК\n{mc['title']}\n{data['name']}\n{message.text}"
    )

    await message.answer("Вы записаны!", reply_markup=bottom_kb())
    await state.clear()

# ================= RUN =================

async def main():
    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())

