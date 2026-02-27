import openpyxl
import os


def load_clubs(file="joined_clubs.xlsx"):
    if not os.path.exists(file):
        return []

    wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
    sheet = wb.active

    headers = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}
        for h, v in zip(headers, row):
            record[h] = str(v).strip() if v else ""
        data.append(record)

    return data